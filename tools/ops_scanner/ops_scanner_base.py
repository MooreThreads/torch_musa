import argparse
import os
import re
from collections import defaultdict
from itertools import chain
from os.path import join

import openpyxl


class OpsScannerBase:
    def __init__(
        self,
        root_dir: str = "./",
        file_ext: str = ".cu",
        ops_regex: str = r"REGISTER_DISPATCH\((.*?)_stub,",
        output_path: str = "ops.xlsx",
        conditional_regex: str = None,
    ):
        self.root_dir = root_dir
        self.file_ext = file_ext
        self.ops_regex = ops_regex
        self.file_paths = []
        self.res = defaultdict(list)
        self.output_path = output_path
        self.conditional_regex = conditional_regex

    def _get_files(self):
        for root, dirs, files in os.walk(self.root_dir):
            for file in files:
                if file.endswith(self.file_ext) and not file.endswith(".md"):
                    self.file_paths.append(join(root, file))

    def _get_op_name(self, file_path: str):
        file_name = os.path.split(file_path)[-1]
        with open(file_path) as f:
            block_match = f.read()
            match_res = []
            if self.conditional_regex:
                block_match = re.search(self.conditional_regex, block_match)
                if block_match:
                    match_res: list = re.findall(self.ops_regex, block_match.group(1))
            else:
                match_res: list = re.findall(self.ops_regex, block_match)
            if match_res:
                processed_res = list(filter(lambda x: x, chain(*match_res)))
                self.res[file_name].extend(processed_res)

    def write_to_xlsx(self):
        # Create a new workbook and select the active worksheet
        workbook = openpyxl.Workbook()
        worksheet = workbook.active

        # Write the headers to the worksheet
        headers = [self.file_ext + " filename", "operator name"]
        for i, header in enumerate(headers):
            cell = worksheet.cell(row=1, column=i + 1)
            cell.value = header

        # Write the values to the worksheet
        res_items = self._convert()
        for i, value in enumerate(res_items):
            for j, header in enumerate(headers):
                cell = worksheet.cell(row=2 + i, column=j + 1)
                cell.value = value[j]

        # Save the workbook
        workbook.save(self.output_path)

    def _convert(self):
        full_items = []
        res_items = list(self.res.items())
        for k, v in res_items:
            for op in v:
                full_items.append((k, op))
        full_items.sort(key=lambda x: x[0])
        return full_items

    def scan(self):
        self._get_files()
        for file_path in self.file_paths:
            self._get_op_name(file_path)


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Scan ops in PyTorch")
    parser.add_argument(
        "--scan-dir", help="Specify directory to be scanned", default="./"
    )
    parser.add_argument("--file-ext", help="Specify file extension", default=".cu")
    parser.add_argument(
        "--regex",
        help="Specify search regular expression",
        default=r"REGISTER_DISPATCH\((.*?)_stub,",
    )
    parser.add_argument(
        "--conditional-regex",
        help="Specify conditional search regular expression",
        default=r"TORCH_LIBRARY_IMPL\(aten, QuantizedPrivateUse1, m\) \{([\s\S]*?)\}",
    )
    parser.add_argument(
        "--output-path",
        help="Specify ops' scanning result location",
        default="./ops.xlsx",
    )

    args = parser.parse_args()
    test_dir = args.scan_dir
    test_file_ext = args.file_ext
    test_ops_regex = args.regex
    test_conditional_regex = args.conditional_regex
    ops_scanner = OpsScannerBase(
        test_dir,
        test_file_ext,
        test_ops_regex,
        output_path=args.output_path,
        condition_regex=test_conditional_regex,
    )
    ops_scanner.scan()
    ops_scanner.write_to_xlsx()
