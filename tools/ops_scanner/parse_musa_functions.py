import argparse
import os
import re
from collections import defaultdict
from itertools import chain
from os.path import join, dirname, abspath

import openpyxl

DEFAULT_MUSA_FUNCTIONS_YAML_PATH = join(
    dirname(dirname(dirname(abspath(__file__)))),
    "torch_musa/csrc/aten/ops/musa_functions.yaml",
)


class MusaFunctionsParser:
    def __init__(
        self,
        ops_regex: str = r"- func: (.*)",
        input_path: str = DEFAULT_MUSA_FUNCTIONS_YAML_PATH,
        output_path: str = "musa_ops.xlsx",
    ):
        self.ops_regex = ops_regex
        self.input_path = input_path
        self.output_path = output_path
        self.res = []

    def parse_file(self):
        with open(self.input_path) as f:
            self.res: list = re.findall(self.ops_regex, f.read(), re.M)

    def write_to_xlsx(self):
        # Create a new workbook and select the active worksheet
        workbook = openpyxl.Workbook()
        worksheet = workbook.active

        # Write the headers to the worksheet
        headers = [
            "Pytorch算子",
            "FP32",
            "FP16",
            "BF16",
            "INT32",
            "INT16",
            "INT8",
            "BOOL",
            "算子类型",
        ]
        for i, header in enumerate(headers):
            cell = worksheet.cell(row=1, column=i + 1)
            cell.value = header

        # Write the values to the worksheet
        res_items = self._convert()
        for i, value in enumerate(res_items):
            for j, header in enumerate(headers):
                cell = worksheet.cell(row=2 + i, column=j + 1)
                cell.value = value[j]

        # Save the workbook
        workbook.save(self.output_path)

    def _convert(self):
        full_items = []
        for op in self.res:
            full_items.append((op, "", "", "", "", "", "", "", ""))
        full_items.sort(key=lambda x: x[0])
        return full_items


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Scan ops from musa_functions.yaml")
    parser.add_argument(
        "--input-path",
        help="Specify yaml file path to be scanned",
        default=DEFAULT_MUSA_FUNCTIONS_YAML_PATH,
    )
    parser.add_argument(
        "--ops-regex", help="Specify search regular expression", default=r"- func: (.*)"
    )
    parser.add_argument(
        "--output-path",
        help="Specify ops' scanning result location",
        default="./musa_ops.xlsx",
    )

    args = parser.parse_args()
    test_ops_regex = args.ops_regex
    test_input_path = args.input_path
    test_output_path = args.output_path
    musa_functions_parser = MusaFunctionsParser(
        test_ops_regex, test_input_path, test_output_path
    )
    musa_functions_parser.parse_file()
    musa_functions_parser.write_to_xlsx()
